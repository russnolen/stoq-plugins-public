#   Copyright 2014-2015 PUNCH Cyber Analytics Group
#
#   Licensed under the Apache License, Version 2.0 (the "License");
#   you may not use this file except in compliance with the License.
#   You may obtain a copy of the License at
#
#       http://www.apache.org/licenses/LICENSE-2.0
#
#   Unless required by applicable law or agreed to in writing, software
#   distributed under the License is distributed on an "AS IS" BASIS,
#   WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
#   See the License for the specific language governing permissions and
#   limitations under the License.

"""
Overview
========

Read content from Excel Documents

"""

from io import BytesIO
from openpyxl import load_workbook

from stoq.plugins import StoqReaderPlugin


class ExcelReader(StoqReaderPlugin):

    def __init__(self):
        super().__init__()

    def activate(self, stoq):
        self.stoq = stoq
        super().activate()

    def read(self, payload, **kwargs):
        """
        Read content from Excel Documents

        :param bytes payload : Content of excel file
        :param int header_row : The row that contains header information (optional)

        :returns: Extracted content of payload
        :rtype: dict

        """

        if 'header_row' in kwargs:
            header_row = kwargs['header_row']
        else:
            header_row = False

        results = {}

        payload_object = BytesIO(payload)

        workbook = load_workbook(filename=payload_object)

        for worksheet in workbook.worksheets:

            # Parse the header row for the dict keys, if there is one
            if header_row:
                headers = []
                for column in range(1, worksheet.max_column):
                    cell = worksheet.cell(row=header_row, column=column).value
                    if cell:
                        cell = cell.lower().replace(" ", "_")
                        headers.append(cell)

            # Iterate over each row and grab all of the values
            rows = []

            for row in worksheet.rows:

                content = {}

                for cell in row:
                    try:
                        if header_row:
                            # This excel file has a header row, so let's
                            # create a dict with the header as the key,
                            # and cell as the value.
                            if cell.row <= header_row:
                                break
                            content[headers[cell.col_idx-1]] = cell.value
                        else:
                            # Process the excel file without a header row.
                            # The key value will be that of the column
                            # (A, B...ZZ )
                            content[cell.column] = cell.value
                    except:
                        pass

                # Make sure we have content to save, otherwise just move along
                if bool(content):
                    # Ensure that each value in the row is not None
                    if not all(cell is None for cell in content.values()):
                        rows.append(content)
            if bool(rows):
                results[worksheet.title] = rows 

        return results

